from openpyxl import Workbook, load_workbook

book = load_workbook("salaries.xlsx")
print(book)
sheet = book['Лист1']
sum = 0
for item in range(2, 10):
    sum += sheet['B' + str(item)].value 
    print(sheet['B' + str(item)].value)
print(sum)
print(sum/8)

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")