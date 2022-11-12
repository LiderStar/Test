from ctypes import alignment
from msilib.schema import Font
from posixpath import abspath
import openpyxl
from os.path import join

from openpyxl import load_workbook


class NoData(Exception):
    pass

data_path = join('.','file name')
data_path = abspath(data_path)

wb = load_workbook(filename = data_path, data_only=True, read_only=True)

wsn = list(wb.sheetnames) # список листов нашего файла

wsdata = None

for i in wsn:
    if wb[i]['B1'].value == "Manager":
        wsdata = i

if wsdata == None:
    raise NoData("Нет необходимого заголовка")

ws = wb[wsdata] # выбирам лист где есть поле Менеджер

title_sheet = [cell.value for cell in next(ws.iter_rows(min_row=1, min_col=1, max_row=1, max_col=ws.max_column))]


#Создаём словарь +
mandata = {}

for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
    if len(row) > 0:
        fio = row[11].value
        if fio is not None:
            fiodata = [cell.value for cell in row]
            if fio not in mandata:
                mandata[fio] = []
            mandata[fio].append(fiodata)

#Вводим полученный списк +
for fio in mandata:
    print(f'Менеджер {fio}, количество : {len(mandata[fio])}')

#Закрыть исходный файл +
wb.close



for i in range(1, 52):
        zagl = ws.cell(row=1, colum=i)
        zagl.alignment = alignment(horizontal='center')
        zagl.font = Font(bold=True, italic=True, color='ffffff', size=14)

#Создаём отчёты по отступлениям

for fio in mandata:
    exname, *_ = fio.split()
    wb = Workbook()
    ws = wb.active
    ws.title = "Manager"
#
    ws.append(column_headings)
    for row in mandata[fio]:
        ws.append(row)
#Форматирование таблицы
nmrow = len(mandata[fio])
for i in range(2, nmrow + 2):
    ws.cell(row=i, colum=1).number_format = '# ##0'
    ws.cell(row=i, colum=2).number_format = '# ##0'
    ws.cell(row=i, colum=3).number_format = '# ##0'
    ws.cell(row=i, colum=4).number_format = '# ##0'
    ws.cell(row=i, colum=5).number_format = '# ##0'
    ws.cell(row=i, colum=6).number_format = '# ##0'
    ws.cell(row=i, colum=7).number_format = '# ##0'
    ws.cell(row=i, colum=8).number_format = '# ##0'
    ws.cell(row=i, colum=9).number_format = '# ##0'
    ws.cell(row=i, colum=10).number_format = '# ##0'
    ws.cell(row=i, colum=11).number_format = '# ##0'
    ws.cell(row=i, colum=12).number_format = '# ##0'
    ws.cell(row=i, colum=13).number_format = '# ##0'
    ws.cell(row=i, colum=14).number_format = '# ##0'
    ws.cell(row=i, colum=15).number_format = '# ##0'
    ws.cell(row=i, colum=16).number_format = '# ##0'
    ws.cell(row=i, colum=17).number_format = '# ##0'
    ws.cell(row=i, colum=18).number_format = '# ##0'
    ws.cell(row=i, colum=19).number_format = '# ##0'
    ws.cell(row=i, colum=20).number_format = '# ##0'
    ws.cell(row=i, colum=21).number_format = '# ##0'
    ws.cell(row=i, colum=22).number_format = '# ##0'
    ws.cell(row=i, colum=23).number_format = '# ##0'
    ws.cell(row=i, colum=24).number_format = '# ##0'
    ws.cell(row=i, colum=25).number_format = '# ##0'
    ws.cell(row=i, colum=26).number_format = '# ##0'

#Устанавливаем ширину столбцов
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['P'].width = 20
    ws.column_dimensions['Q'].width = 20
    ws.column_dimensions['R'].width = 20
    ws.column_dimensions['S'].width = 20
    ws.column_dimensions['T'].width = 20
    ws.column_dimensions['U'].width = 20
    ws.column_dimensions['V'].width = 20
    ws.column_dimensions['W'].width = 20
    ws.column_dimensions['X'].width = 20
    ws.column_dimensions['Y'].width = 20
    ws.column_dimensions['Z'].width = 20

#Собираем путь где сохранить файл
exfilname = join('.', 'Data', (exname + '.xlsx'))
exfilname = abspatn(exfilname)
wb.close

wb.save(exfilname)
wb.close

print ('\nВсе данные из исходного файла обработаны.')
print ('Файлы сформированы и сохранены в каталог')